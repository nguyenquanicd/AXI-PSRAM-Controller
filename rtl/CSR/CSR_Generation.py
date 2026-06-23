#--------------------------------------------------------------------
# Project: APB-CSR-Generator
# Author: Trthinh (Ethan)
# Function: CSR Generation Tool
# Page: VLSI Technology
#---------------------------------------

import sys
import openpyxl
import getpass
from datetime import datetime
import os
import shutil
from sys import argv

RSVD_KEY = 'reserved'
def bit_width(field):
    if ":" in field:
        temp = field.replace("[","")
        temp = temp.replace("]","")
        temp = temp.split(':')
        return int(temp[0]) - int(temp[1]) + 1
    else:
        return 1
class port:
    def __init__(self,name,direction, width):
        self.name = name
        self.direction = direction
        self.width = width
        self.rtl = ''
        self.port_declaration()
    def port_declaration(self):
        if self.width > 1:
            self.rtl = "  " + str(self.direction) + " [" + str(self.width-1) + ":0] " + self.name + ",\n"
        elif self.width < 1:
            print(f"ERROR: The width of port {self.name} is incorrect")
            sys.exit()
        else:
            self.rtl = "  " + str(self.direction) + " " + self.name + ",\n"
        if self.direction == 'logic':
            self.rtl = self.rtl.replace(',',';')
class register:
    def __init__(self,name,offset,reset,description,bit_list):
        self.name = name
        self.offset = offset
        self.reset = reset
        self.description = description
        self.bit = bit_list
        self.port_list = []
        self.logic_list = []
        self.we_rtl = ''
        self.gen_port_logic()
        self.lcparam = f"  localparam PARA_{self.name.upper()}_OFFSET = 16'h{self.offset[2:]},\n"
        self.rtl = ''
        self.gen_rtl()
    # def add_1_bit(self,bit):
    #     self.bit.append(bit)
    # def add_multi_bit(self,list):
    #     self.bit += list   
    def gen_port_logic(self):
        we_reg = 0
        for b in self.bit:
            if b.type == 'ro' and b.name != RSVD_KEY:
                self.port_list.append (port(f"i_{self.name}_{b.name}","input",b.width))
            else:
                if we_reg == 0:
                    self.we_rtl = f"  assign we_{self.name} = apb_write_en & (reg_address == PARA_{self.name.upper()}_OFFSET);\n"
                    self.logic_list.append (port(f"we_{self.name}","logic",1))
                    we_reg = 1
                if b.type == 'rw':
                    self.port_list.append (port(f"o_{self.name}_{b.name}","output",b.width))
                    self.logic_list.append (port(f"reg_{self.name}_{b.name}","logic",b.width))
                    self.logic_list.append (port(f"nxt_{self.name}_{b.name}","logic",b.width))
                elif b.type == 'rwi' or b.type == 'w1c':
                    self.port_list.append (port(f"o_{self.name}_{b.name}","output",b.width))
                    self.port_list.append (port(f"i_hw_wdata_{self.name}_{b.name}","input",b.width))
                    self.port_list.append (port(f"i_hw_we_{self.name}_{b.name}","input",1))
                    self.logic_list.append (port(f"reg_{self.name}_{b.name}","logic",b.width))
                    self.logic_list.append (port(f"nxt_{self.name}_{b.name}","logic",b.width))
            
        self.logic_list.append (port(f"{self.name}_value","logic",32))
    def gen_rtl(self):
        temp = ""
        for b in self.bit:
            b.next_value_cal()
            b.ff_assignment()
            self.rtl += b.rtl
            if b.type == 'ro':
                if b.name == RSVD_KEY:
                    temp += f"    {self.name}_value{b.field} = '0;\n"
                else:
                    temp += f"    {self.name}_value{b.field} = i_{self.name}_{b.name};\n"
            else:
                temp += f"    {self.name}_value{b.field} = reg_{self.name}_{b.name};\n"
        self.rtl += f'''
  //--------------------------------------------------------------------
  //Combine the value of {self.name}
  //--------------------------------------------------------------------
  always_comb begin
{temp[:-1]}
  end'''

class bit:
    def __init__(self,name,field,type,reset_val,description, reg_group, async_option):
        self.name = name
        self.field = field
        self.type = type
        self.reset_val = reset_val
        self.description = description
        self.port_list = []
        self.logic_list = []
        if int(async_option):
            self.clock = 'i_reg_clk'
            self.reset = 'i_reg_rstn'
        else:
            self.clock = 'i_bus_clk'
            self.reset = 'i_bus_rstn'
        self.width = bit_width(self.field)
        self.reg_group = reg_group
        self.rtl = ''
    def ff_assignment(self):
        register = self.reg_group
        reset_val_short = self.reset_val.replace('0x',"")
        if self.type != 'ro':
            self.rtl += f'''  //FF for bit {register}_{self.name}
  always_ff @ (posedge {self.clock}, negedge {self.reset}) begin //FF for each bit
    if(!{self.reset}) 
      reg_{register}_{self.name} <= {self.width}'h{reset_val_short};
    else
      reg_{register}_{self.name} <= nxt_{register}_{self.name};
  end 
  assign o_{register}_{self.name} = reg_{register}_{self.name};\n'''

    def next_value_cal(self):
        space_num = 13 + len(self.reg_group) + 1 + len(self.name) + 1
        if self.type == 'rw':
            self.rtl += f'''  //Assignment for next value of {self.reg_group}_{self.name} rw
  assign nxt_{self.reg_group}_{self.name} = (we_{self.reg_group}) ? reg_pwdata{self.field} : reg_{self.reg_group}_{self.name};\n'''
        elif self.type == 'rwi':
            self.rtl += f'''  //Assignment for next value of {self.reg_group}_{self.name} rwi
  assign nxt_{self.reg_group}_{self.name} = (we_{self.reg_group}) ? reg_pwdata{self.field}
{" " * space_num}: (i_hw_we_{self.reg_group}_{self.name}) ? i_hw_wdata_{self.reg_group}_{self.name}
{" " * space_num}: reg_{self.reg_group}_{self.name};\n'''
        elif self.type == 'w1c':
            self.rtl += f'''  //Assignment for next value of {self.reg_group}_{self.name} w1c
  assign nxt_{self.reg_group}_{self.name} = (we_{self.reg_group}) ? (~reg_pwdata{self.field} & reg_{self.reg_group}_{self.name})
{" " * space_num}: (i_hw_we_{self.reg_group}_{self.name}) ? i_hw_wdata_{self.reg_group}_{self.name}
{" " * space_num}: reg_{self.reg_group}_{self.name};\n'''

class apb_protocol:
    def __init__(self, data_wd ,addr_wd, async_option):
        self.data_wd = data_wd
        self.addr_wd = addr_wd
        self.async_option = async_option
        self.port_list = []
        self.logic_list = []
        self.gen_port()
        self.gen_logic()
        self.apb_access()
        self.rw_rtl()
    def gen_port(self):
        self.port_list.append (port("i_bus_clk","input",1))
        self.port_list.append (port("i_bus_rstn","input",1))
        self.port_list.append (port("i_paddr","input",self.addr_wd))
        self.port_list.append (port("i_protect_en","input",1))
        self.port_list.append (port("i_slverr_en","input",1))
        self.port_list.append (port("i_pprot","input",3))
        self.port_list.append (port("i_pwdata","input",self.data_wd))
        self.port_list.append (port("i_pwrite","input",1))
        self.port_list.append (port("i_penable","input",1))
        self.port_list.append (port("i_psel","input",1))
        self.port_list.append (port("i_pstrb","input",4))
        self.port_list.append (port("o_pslverr","output",1))
        self.port_list.append (port("o_pready","output",1))
        self.port_list.append (port("o_prdata","output",self.data_wd))
        if int(self.async_option):
            self.port_list.append (port("i_reg_clk","input",1))
            self.port_list.append (port("i_reg_rstn","input",1))
    def gen_logic(self):
        self.logic_list.append (port("apb_setup","logic",1))
        self.logic_list.append (port("apb_protect","logic",1))
        self.logic_list.append (port("apb_slverr","logic",1))
        self.logic_list.append (port("apb_complete","logic",1))
        self.logic_list.append (port("apb_write","logic",1))
        self.logic_list.append (port("apb_read","logic",1))

        self.logic_list.append (port("reg_address","logic",self.addr_wd))
        self.logic_list.append (port("reg_pwdata","logic",self.data_wd))
        self.logic_list.append (port("reg_slverr","logic",1))

        self.logic_list.append (port("apb_read_en","logic",1))
        self.logic_list.append (port("apb_write_en","logic",1))

        self.logic_list.append (port("nxt_prdata","logic",self.data_wd))
        self.logic_list.append (port("reg_pready","logic",1))
        self.logic_list.append (port("reg_prdata","logic",self.data_wd))
        
        if int(self.async_option):
            self.logic_list.append (port("read_ack","logic",1))
            self.logic_list.append (port("read_ack_bus_clk","logic",1))
            self.logic_list.append (port("reg_read_req","logic",1))
            self.logic_list.append (port("reg_read_ack_reg_clk_dly","logic",1))
            self.logic_list.append (port("read_aclk_bus_clk_falling","logic",1))

            self.logic_list.append (port("write_ack","logic",1))
            self.logic_list.append (port("write_ack_bus_clk","logic",1))
            self.logic_list.append (port("reg_write_req","logic",1))
            self.logic_list.append (port("reg_write_ack_reg_clk_dly","logic",1))
            self.logic_list.append (port("write_aclk_bus_clk_falling","logic",1))

            self.logic_list.append (port("read_req_reg_clk","logic",1))
            self.logic_list.append (port("reg_read_req_reg_clk_dly","logic",1))

            self.logic_list.append (port("write_req_reg_clk","logic",1))
            self.logic_list.append (port("reg_write_req_reg_clk_dly","logic",1))
            self.logic_list.append (port("reg_read_ack_bus_clk_dly","logic",1))
            self.logic_list.append (port("reg_write_ack_bus_clk_dly","logic",1))
        else:
            self.logic_list.append (port("reg_apb_read_capture","logic",1))
            self.logic_list.append (port("reg_apb_write_capture","logic",1))
            

    def gen_port_rtl(self):
        self.port_rtl = ''
        for p in self.port_list:
            self.port_rtl += p.rtl

    def apb_access(self):

        self.access = """
  //APB general access phase assignment - START
  assign apb_setup = i_psel & ~i_penable;
  assign apb_protect = i_protect_en ? ~i_pprot[1] : 1'b1;
  assign apb_slverr = i_slverr_en ? (~apb_protect      //error in protection
                      | (|i_paddr[1:0]) //error in address 
                      | (~&i_pstrb & i_pwrite)) : 1'b0;    //error in pstrb signal
  assign apb_complete = apb_setup & (~apb_slverr);
  assign apb_write = i_pwrite & apb_complete;
  assign apb_read = ~i_pwrite & apb_complete; 
  
  //APB capture FF phase
  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin //address capture FF
    if(!i_bus_rstn) 
      reg_address <= '0;
    else if (apb_complete)
      reg_address <= i_paddr;
  end
  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin //write data capture FF
    if(!i_bus_rstn) 
      reg_pwdata <= '0;
    else if (apb_write)
      reg_pwdata <= i_pwdata;
  end
  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin //slverr output FF
    if(!i_bus_rstn) 
      reg_slverr <= '0;
    else if (apb_setup & i_slverr_en)
      reg_slverr <= apb_slverr;
  end
  assign o_pslverr = reg_slverr;
  //--------------------------------------------------------------------"""
    
    def rw_rtl(self):
        if int(self.async_option):
            self.rtl_rw = '''
  //APB handshake synchronization (appear when option Async is selected) - START
  //--------------------------------------------------------------------
  //Bus clock - reading phase
  //--------------------------------------------------------------------
  m_vlsi_synch #(
    .PARA_LEVELS (2)
  ) u_read_ack_reg2bus (
    .i_clk (i_bus_clk),
    .i_rstn (i_bus_rstn),
    .i_data_in (read_ack),
    .o_data_out (read_ack_bus_clk)
  );
  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin
    if(!i_bus_rstn) 
      reg_read_req <= '0;
    else begin
      casez ({read_ack_bus_clk, apb_read})
        2'b1?: reg_read_req <= '0;
        2'b01: reg_read_req <= '1;
        2'b00: reg_read_req <= reg_read_req;
        default: reg_read_req <= 'x;
      endcase
    end
  end
  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin
    if(!i_bus_rstn) 
      reg_read_ack_bus_clk_dly <= '0;
    else
      reg_read_ack_bus_clk_dly <= read_ack_bus_clk;
  end
  assign read_aclk_bus_clk_falling = reg_read_ack_bus_clk_dly & ~read_ack_bus_clk;

  //--------------------------------------------------------------------
  //Bus clock - writing phase
  //--------------------------------------------------------------------
  m_vlsi_synch #(
    .PARA_LEVELS (2)
  ) u_write_ack_reg2bus (
    .i_clk (i_bus_clk),
    .i_rstn (i_bus_rstn),
    .i_data_in (write_ack),
    .o_data_out (write_ack_bus_clk)
  );
  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin
    if(!i_bus_rstn) 
      reg_write_req <= '0;
    else begin
      casez ({write_ack_bus_clk, apb_write})
        2'b1?: reg_write_req <= '0;
        2'b01: reg_write_req <= '1;
        2'b00: reg_write_req <= reg_write_req;
        default: reg_write_req <= 'x;
      endcase
    end
  end
  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin
    if(!i_bus_rstn) 
      reg_write_ack_bus_clk_dly <= '0;
    else
      reg_write_ack_bus_clk_dly <= write_ack_bus_clk;
  end
  assign write_aclk_bus_clk_falling = reg_write_ack_bus_clk_dly & ~write_ack_bus_clk;

  //--------------------------------------------------------------------
  //Reg clock - reading handshake
  //--------------------------------------------------------------------
  m_vlsi_synch #(
    .PARA_LEVELS (2)
  ) u_rd_ack_reg2bus (
    .i_clk (i_reg_clk),
    .i_rstn (i_reg_rstn),
    .i_data_in (reg_read_req),
    .o_data_out (read_req_reg_clk)
  );
  always_ff @ (posedge i_reg_clk, negedge i_reg_rstn) begin
    if(!i_reg_rstn) 
      reg_read_req_reg_clk_dly <= '0;
    else
      reg_read_req_reg_clk_dly <= read_req_reg_clk;
  end
  assign apb_read_en = ~reg_read_req_reg_clk_dly & read_req_reg_clk;
  assign read_ack = reg_read_req_reg_clk_dly;

  //--------------------------------------------------------------------
  //Reg clock - writing handshake
  //--------------------------------------------------------------------
  m_vlsi_synch #(
    .PARA_LEVELS (2)
  ) u_wr_ack_reg2bus (
    .i_clk (i_reg_clk),
    .i_rstn (i_reg_rstn),
    .i_data_in (reg_write_req),
    .o_data_out (write_req_reg_clk)
  );
  always_ff @ (posedge i_reg_clk, negedge i_reg_rstn) begin
    if(!i_reg_rstn) 
      reg_write_req_reg_clk_dly <= '0;
    else
      reg_write_req_reg_clk_dly <= write_req_reg_clk;
  end
  assign apb_write_en = ~reg_write_req_reg_clk_dly & write_req_reg_clk;
  assign write_ack = reg_write_req_reg_clk_dly;

  //--------------------------------------------------------------------
  //PREADY phase
  //--------------------------------------------------------------------
  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin
    if(!i_bus_rstn) 
      reg_pready <= '0;
    else
      reg_pready <= read_aclk_bus_clk_falling // reading complete
                | write_aclk_bus_clk_falling // writing complete
                | (apb_slverr & apb_setup);//slverr assert
  end
  assign o_pready = reg_pready;
  //--------------------------------------------------------------------'''
        else:
            self.rtl_rw = '''
  //APB read/write register (appear when option Async is not selected) - START
  //--------------------------------------------------------------------
  //Reading phase
  //--------------------------------------------------------------------
  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin
    if(!i_bus_rstn) 
      reg_apb_read_capture <= '0;
    else
      reg_apb_read_capture <= apb_read;
  end
  assign apb_read_en = reg_apb_read_capture;

  //--------------------------------------------------------------------
  //Writing phase
  //--------------------------------------------------------------------
  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin
    if(!i_bus_rstn) 
      reg_apb_write_capture <= '0;
    else
      reg_apb_write_capture <= apb_write;
  end
  assign apb_write_en = reg_apb_write_capture;

  //--------------------------------------------------------------------
  //PREADY phase
  //--------------------------------------------------------------------
  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin
    if(!i_bus_rstn) 
      reg_pready <= '0;
    else
      reg_pready <= apb_write_en // reading complete
                | apb_read_en // writing complete
                | (apb_slverr & apb_setup);//slverr assert
  end
  assign o_pready = reg_pready;
  
  //--------------------------------------------------------------------'''
    def prdata_rtl(self, reg_list):
        rtl_prdata = ""
        decoder = '''  //--------------------------------------------------------------------
  //O_PRDATA phase
  //--------------------------------------------------------------------
  always_comb begin
    case (reg_address)
      default nxt_prdata = '0;
    endcase
  end\n'''     
        flag = "      default nxt_prdata = '0;"
        space = "      "
        for r in reg_list:
            decoder = decoder.replace(flag, space + "PARA_" + r.name.upper() + "_OFFSET: nxt_prdata = " + r.name + "_value;\n" + flag) 
            
        rtl_prdata += decoder
        rtl_prdata += '''  always_ff @ (posedge i_bus_clk, negedge i_bus_rstn) begin
    if(!i_bus_rstn) 
      reg_prdata <= '0;
    else if (apb_read_en)
      reg_prdata <= nxt_prdata;
  end
  assign o_prdata = reg_prdata;\n'''
        self.rtl_prdata = rtl_prdata

class sfr:
    def __init__(self,module_name, addr_wd, data_wd, reg_list, protocol, async_option):
        self.module_name = module_name
        self.addr_wd = addr_wd
        self.data_wd = data_wd
        self.async_option = async_option
        if protocol.lower() == 'apb':
            self.protocol = apb_protocol(data_wd=data_wd,
                                         addr_wd=addr_wd,
                                         async_option=async_option)
        
        self.rtl = ''
        self.reg_list = reg_list
        self.gen_port_logic_rtl()
        self.port_rtl = self.port_rtl[:-2]
        self.logic_rtl = self.logic_rtl[:-1]
    def gen_port_logic_rtl(self):
        self.port_rtl = ''
        self.logic_rtl = ''
        for p in self.protocol.port_list:
            self.port_rtl += p.rtl
        for l in self.protocol.logic_list:
            self.logic_rtl += l.rtl
        for r in self.reg_list:
            for pr in r.port_list:
                self.port_rtl += pr.rtl
            for lr in r.logic_list:
                self.logic_rtl += lr.rtl
    def gen_json(self):
        reg_json = ''
        bit_json = ''
        bracket_o = "{"
        bracket_c = "}"
        for r in self.reg_list:
            reg_json += f'''        {bracket_o}
        "Register Name": "{r.name}",
        "Offset": "{r.offset}",
        "Reset Value": "{r.reset}",
        "Descriprtion": "{r.description}",
        "Bits": ['''
            for b in r.bit:
                bit_json += f'''
            {bracket_o}
                "Bit Name": "{b.name}",
                "Bit Type": "{b.type}",
                "Bit Field": "{b.field}",
                "Reset Value": "{b.reset_val}",
                "Description": "{b.description}"
            {bracket_c},'''
            reg_json += bit_json[:-1] + f"\n{' ' * 10}]\n" + f"{' ' * 8}{bracket_c},\n"


        reg_json = reg_json[:-2]
        self.json = f'''    {bracket_o}
    "Project": "APB-CSR-Generator",
    "Module": "JSON debug file",
    "Function": "Used to check the configuration after generation",
    "Author": "Trthinh (Ethan)",
    "Generator": "{getpass.getuser()}",
    "Page": "VLSI Technology", 
    "Table - Configuration": [
    {bracket_o}
        "Module Name": "{self.module_name}",
        "Address Width": "{self.addr_wd}",
        "Data Width": "{self.data_wd}",
        "Protocol": "APB",
        "Asynchornous": "{self.async_option}"
    {bracket_c}
    ],
    "Table - Register Definition": [
{reg_json}
    ]
{bracket_c}'''    
    def gen_rtl(self):
        header = f'''`timescale 1ns/1ps
//--------------------------------------------------------------------
// Project: APB-CSR-Generator
// Generator: {getpass.getuser()}
// Date and time: {str(datetime.now())}
// Module: {self.module_name}
// Function: Control & Status register via APB interface
// Page: VLSI Technology
//--------------------------------------------------------------------'''
        localparam_offset = ''
        reg_rtl = ''
        reg_we = ''
        for r in self.reg_list:
            localparam_offset += r.lcparam
            reg_rtl += r.rtl
            reg_we += r.we_rtl
        localparam_offset = localparam_offset[:-2] + "\n"
        self.protocol.prdata_rtl(self.reg_list)
        body = f'''
module {self.module_name} # (
{localparam_offset}
)(
{self.port_rtl}
);
  //Logic signal declaration
{self.logic_rtl}
  //end of logic signal declaration
{self.protocol.access}
{reg_we}
{self.protocol.rtl_rw}

{reg_rtl}
{self.protocol.rtl_prdata}
endmodule: {self.module_name}'''
        self.rtl = header + body
    def gen_tb(self):
        logic = self.port_rtl.replace (' input ', ' logic ')
        logic = logic.replace (' output ', ' logic ')
        logic = logic.replace(",",";")
        logic = logic[:-1] + ";\n"
        add_ins = logic.replace(" logic ","")
        add_ins = add_ins.replace(";",",")
        tb = f'''module tb;
{logic}
{self.module_name} u_dut (
{add_ins[:-1]});'''
        self.tb = tb

# MAIN FUNCTION
print("Start the CSR Generation!")
print("The register configured will be shown below!")
current_path = os.getcwd()
script_path = os.path.dirname(os.path.realpath(__file__))
wb = openpyxl.load_workbook(argv[1])
sheet_name = argv[2]
ws = wb[sheet_name]
start_config = 0
start_reg = 0
reg_list = []
start_col = 0 # Can be updated later
reg_col = start_col
offset_col = start_col + 1
reset_col = start_col + 5
des_col = start_col + 6
bit_name_col = start_col + 2
bit_field_col = start_col + 3
bit_type_col = start_col + 4

reg_name = ''
for row in range(1,ws.max_row + 1):
    if start_reg == 0:
        if start_config == 0:
            if 'Table - Configuration' in str(ws[row][start_col].value):
                start_config = 1
        else:
            if 'Module Name' in str(ws[row][start_col].value):
                module_name = ws[row][start_col + 1].value
            if 'Protocol' in str(ws[row][start_col].value):
                protocol = ws[row][start_col + 1].value
            if 'Data Width' in str(ws[row][start_col].value):
                data_wd = int(ws[row][start_col + 1].value)
            if 'Address Width' in str(ws[row][start_col].value):
                addr_wd = int(ws[row][start_col + 1].value)
            if 'Asynchronous' in str(ws[row][start_col].value):
               async_option = int(ws[row][start_col + 1].value)
            if 'REGISTER' in str(ws[row][start_col].value):
               start_reg = 1
    else:
        if ws[row][0].value is not None:
            if reg_name != '':
                r = register(name=reg_name,
                            offset=offset,
                            reset=reset_val,
                            description=description,
                            bit_list=bit_list)
                reg_list.append(r)
            reg_name = ws[row][reg_col].value
            offset = ws[row][offset_col].value
            reset_val = ws[row][reset_col].value
            description = ws[row][des_col].value
            bit_list = []
            
            
        else:
            bit_name = ws[row][bit_name_col].value
            bit_field = ws[row][bit_field_col].value
            bit_type = ws[row][bit_type_col].value
            bit_reset = ws[row][reset_col].value
            bit_des = ws[row][des_col].value
            print("Register: ", reg_name,", Bit name: ", bit_name, ", Bit Type: ", bit_type)
            b = bit(name=bit_name,
                    field=bit_field,
                    type=bit_type,
                    reset_val=bit_reset,
                    description=bit_des,
                    reg_group=reg_name,
                    async_option=async_option)
            bit_list.append(b)
r = register(name=reg_name,
            offset=offset,
            reset=reset_val,
            description=description,
            bit_list=bit_list)
reg_list.append(r)
s = sfr(module_name, addr_wd, data_wd, reg_list, protocol, async_option)
s.gen_rtl()

print("The current path: ", current_path)
print("The tool path: ", script_path)

rtl_path = current_path + "/RTL"

os.makedirs(rtl_path, exist_ok=True)
with open (rtl_path + '/' + s.module_name + '.sv','w') as f:
    f.write(s.rtl)
models_path = rtl_path + '/models'
if os.path.exists(models_path):
    shutil.copyfile(script_path + "/models/m_vlsi_synch.sv", models_path + "/m_vlsi_synch.sv")
else:
    shutil.copytree(script_path + "/models", rtl_path + '/models')
with open (rtl_path + '/filelist.f','w') as f:
    header_fl = f'''//--------------------------------------------------------------------
// Project: APB-CSR-Generator
// Module: RTL File List
// Author: Trthinh (Ethan)
// Function: Collect all RTL of APB-CSR-Generation Tool
// Page: VLSI Technology
//--------------------------------------------------------------------
'''
    f.write(header_fl)
    f.write(rtl_path + '/' + s.module_name + '.sv\n')
    f.write(rtl_path + '/models/m_vlsi_synch.sv')

s.gen_json()
debug_path = current_path + "/Debug"
os.makedirs(debug_path, exist_ok=True)
with open (debug_path + '/debug.json','w') as f:
    f.write(s.json)
            
print("End of the CSR generation! Thank you.")